from __future__ import annotations

"""source_summary task — dataset source 파일 프리뷰(컬럼/행수/샘플) 계산.

Go control-plane ``loadDatasetSourceSummary``(DuckDB + excelize)의 worker 이전
(ADR-024: 파일 스캔·집계는 worker 책임). Node control-plane이 dataset version
목록/상세에서 metadata ``source_summary`` 캐시가 없는 legacy 버전에 대해
``POST /tasks/source_summary``로 호출한다.

응답 shape은 Go ``domain.DatasetSourceSummary``의 JSON marshal과 동일 (omitempty
필드는 키 생략):
  {available, status, format?, row_count?, column_count?,
   columns?: [{name, type?}], sample_limit?, sample_rows?, error_message?}

계약 유지 주의 (Go 대비 값 parity):
- csv/tsv: ``read_csv_auto(HEADER=TRUE)`` — 기본 샘플링 타입추론 (SAMPLE_SIZE=-1 아님).
- jsonl: ``read_json_auto``. parquet: ``read_parquet``.
- xlsx/xlsm: openpyxl 첫 시트. Go readXlsxSourceSummary와 동일하게 헤더 빈 칸 열
  무시, trim 후 전부 빈 행 skip, 값은 문자열, 컬럼 type "VARCHAR".
- sample_rows 값 인코딩(날짜 등)은 Go 캐시 writer와 미세하게 다를 수 있다 —
  현재 소비처(Node version 목록/상세)는 row_count/columns만 쓰므로 영향 없음.
  확인 필요: upload(캐시 생성) 경로를 Node로 포팅할 때 값 인코딩 parity.
"""

import os
from typing import Any

from .obs import get

_LOG = get(__name__)

_FORMAT_BY_SUFFIX = (
    (".parquet", "parquet"),
    (".csv", "csv"),
    (".tsv", "tsv"),
    (".jsonl", "jsonl"),
    (".ndjson", "jsonl"),
    (".xlsx", "xlsx"),
    (".xlsm", "xlsx"),
)


def run_source_summary(payload: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(payload, dict):
        raise ValueError("source_summary payload must be an object")
    storage_uri = str(payload.get("storage_uri") or "").strip()
    sample_limit_raw = payload.get("sample_limit") or 0
    try:
        sample_limit = int(sample_limit_raw)
    except (TypeError, ValueError):
        raise ValueError("source_summary 'sample_limit' must be an integer")
    summary = build_source_summary(storage_uri, sample_limit)
    _LOG.info(
        "source_summary.completed",
        status=summary.get("status"),
        format=summary.get("format"),
        row_count=summary.get("row_count"),
        column_count=summary.get("column_count"),
    )
    return summary


def build_source_summary(storage_uri: str, sample_limit: int) -> dict[str, Any]:
    summary: dict[str, Any] = {"available": False, "status": "unavailable"}
    if sample_limit > 0:
        summary["sample_limit"] = sample_limit
    if not storage_uri:
        summary["status"] = "missing"
        summary["error_message"] = "storage_uri is required"
        return summary

    source_format = _infer_format(storage_uri)
    if not source_format:
        summary["status"] = "unsupported"
        summary["error_message"] = "unsupported source format"
        return summary
    summary["format"] = source_format

    if not os.path.exists(storage_uri):
        summary["status"] = "missing"
        summary["error_message"] = "source file not found"
        return summary
    if os.path.isdir(storage_uri):
        summary["status"] = "error"
        summary["error_message"] = "source path must be a file"
        return summary

    try:
        if source_format == "xlsx":
            columns, row_count, sample_rows = _read_xlsx_summary(storage_uri, sample_limit)
        else:
            columns, row_count, sample_rows = _read_duckdb_summary(
                storage_uri, source_format, sample_limit
            )
    except Exception as exc:  # 파일 파손/파서 오류 — Go도 status=error로 응답
        _LOG.warning(
            "source_summary.read_failed",
            format=source_format,
            error_category=type(exc).__name__,
            error_message=str(exc),
        )
        summary["status"] = "error"
        summary["error_message"] = str(exc)
        return summary

    summary["available"] = True
    summary["status"] = "ready"
    summary["row_count"] = row_count
    if len(columns) > 0:
        summary["column_count"] = len(columns)
        summary["columns"] = columns
    if sample_rows:
        summary["sample_rows"] = sample_rows
    return summary


def _infer_format(path: str) -> str:
    normalized = path.strip().lower()
    for suffix, source_format in _FORMAT_BY_SUFFIX:
        if normalized.endswith(suffix):
            return source_format
    return ""


def _duckdb_relation(path: str, source_format: str) -> str:
    escaped = path.replace("'", "''")
    if source_format == "parquet":
        return f"read_parquet('{escaped}')"
    if source_format in ("csv", "tsv"):
        # Go와 동일 — SAMPLE_SIZE=-1(전체 스캔 타입추론) 없이 기본 샘플링 프리뷰.
        return f"read_csv_auto('{escaped}', HEADER=TRUE)"
    if source_format == "jsonl":
        return f"read_json_auto('{escaped}')"
    raise ValueError("unsupported source format")


def _read_duckdb_summary(
    path: str, source_format: str, sample_limit: int
) -> tuple[list[dict[str, Any]], int, list[dict[str, Any]]]:
    import duckdb  # heavy dep — lazy import (executor와 동일 패턴)

    relation = _duckdb_relation(path, source_format)
    con = duckdb.connect()
    try:
        columns: list[dict[str, Any]] = []
        for row in con.execute(f"DESCRIBE SELECT * FROM {relation}").fetchall():
            name = str(row[0] or "").strip()
            if not name:
                continue
            column: dict[str, Any] = {"name": name}
            column_type = str(row[1] or "").strip()
            if column_type:
                column["type"] = column_type
            columns.append(column)

        row_count = int(con.execute(f"SELECT COUNT(*) FROM {relation}").fetchone()[0])

        sample_rows: list[dict[str, Any]] = []
        if sample_limit > 0:
            cursor = con.execute(f"SELECT * FROM {relation} LIMIT {sample_limit}")
            names = [desc[0] for desc in cursor.description]
            for values in cursor.fetchall():
                sample_rows.append(
                    {name: _sample_json_value(value) for name, value in zip(names, values)}
                )
        return columns, row_count, sample_rows
    finally:
        con.close()


def _sample_json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (bytes, bytearray)):
        return value.decode("utf-8", errors="replace")
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _read_xlsx_summary(
    path: str, sample_limit: int
) -> tuple[list[dict[str, Any]], int, list[dict[str, Any]]]:
    """Go readXlsxSourceSummary와 동일 semantics — 첫 시트, 첫 행 헤더(빈 칸 열 무시),
    trim 후 전부 빈 행 skip, 셀 값은 문자열, 컬럼 type "VARCHAR"."""
    from openpyxl import load_workbook  # heavy optional dep — lazy import

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        if worksheet is None:
            return [], 0, []
        row_iter = worksheet.iter_rows(values_only=True)
        try:
            header_cells = next(row_iter)
        except StopIteration:
            return [], 0, []

        column_indexes: list[int] = []
        columns: list[dict[str, Any]] = []
        for index, raw in enumerate(header_cells):
            name = ("" if raw is None else str(raw)).strip()
            if not name:
                continue
            column_indexes.append(index)
            columns.append({"name": name, "type": "VARCHAR"})

        row_count = 0
        sample_rows: list[dict[str, Any]] = []
        for cells in row_iter:
            texts = ["" if cell is None else str(cell) for cell in cells]
            if all(not text.strip() for text in texts):
                continue
            row_count += 1
            if sample_limit > 0 and len(sample_rows) < sample_limit:
                record: dict[str, Any] = {}
                for position, index in enumerate(column_indexes):
                    record[columns[position]["name"]] = texts[index] if index < len(texts) else ""
                sample_rows.append(record)
        return columns, row_count, sample_rows
    finally:
        workbook.close()
