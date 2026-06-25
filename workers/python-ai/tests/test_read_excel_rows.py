"""엑셀(.xlsx) 업로드 파싱 — runtime.common._read_excel_rows / _iter_rows 디스패치.

엑셀 업로드를 CSV와 동일한 형태(헤더 dict, 값 문자열)로 읽어 clean/schema_inference가
같은 경로를 타게 한다 (silverone 2026-06-24, 엑셀 업로드 지원).
"""

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from python_ai_worker.runtime import common as rt


def _write_xlsx(path: Path, header: list, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(str(path))


class ReadExcelRowsTest(unittest.TestCase):
    def setUp(self) -> None:
        self._dir = Path(tempfile.mkdtemp())

    def test_basic_header_and_values_as_strings(self):
        path = self._dir / "data.xlsx"
        _write_xlsx(path, ["제목", "좋아요수", "본문"], [
            ["축제 후기", 123, "맥주가 맛있었다"],
            ["둘째날", 0, "사람이 많았다"],
        ])
        rows = rt._read_excel_rows(path)
        self.assertEqual(len(rows), 2)
        # CSV(DictReader)처럼 값은 전부 문자열
        self.assertEqual(rows[0], {"제목": "축제 후기", "좋아요수": "123", "본문": "맥주가 맛있었다"})
        self.assertEqual(rows[1]["좋아요수"], "0")

    def test_empty_cells_become_empty_string_and_empty_rows_skipped(self):
        path = self._dir / "gaps.xlsx"
        _write_xlsx(path, ["a", "b"], [
            ["x", None],
            [None, None],   # 완전히 빈 행 → skip
            [None, "y"],
        ])
        rows = rt._read_excel_rows(path)
        self.assertEqual(rows, [{"a": "x", "b": ""}, {"a": "", "b": "y"}])

    def test_blank_header_column_ignored(self):
        path = self._dir / "blankhdr.xlsx"
        _write_xlsx(path, ["keep", None], [["v1", "drop"]])
        rows = rt._read_excel_rows(path)
        self.assertEqual(rows, [{"keep": "v1"}])

    def test_header_only_returns_empty(self):
        path = self._dir / "headeronly.xlsx"
        _write_xlsx(path, ["a", "b"], [])
        self.assertEqual(rt._read_excel_rows(path), [])

    def test_iter_rows_dispatches_xlsx(self):
        path = self._dir / "dispatch.xlsx"
        _write_xlsx(path, ["text"], [["행1"], ["행2"]])
        rows = rt._iter_rows(str(path))
        self.assertEqual([r["text"] for r in rows], ["행1", "행2"])

    def test_iter_rows_rejects_unknown_extension(self):
        path = self._dir / "data.bin"
        path.write_text("x", encoding="utf-8")
        with self.assertRaises(ValueError):
            rt._iter_rows(str(path))


if __name__ == "__main__":
    unittest.main()
